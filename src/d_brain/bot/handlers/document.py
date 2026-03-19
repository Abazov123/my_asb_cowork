"""Document message handler (.txt, .md, .pdf, .xlsx, .docx)."""

import io
import logging
from datetime import datetime

from aiogram import Bot, Router
from aiogram.types import Message

from d_brain.config import get_settings
from d_brain.services.session import SessionStore
from d_brain.services.storage import VaultStorage

router = Router(name="document")
logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".xlsx", ".docx"}
SUPPORTED_MIMES = {
    "text/plain",
    "text/markdown",
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}
MAX_TEXT_CHARS = 50_000


def _detect_extension(filename: str | None, mime: str | None) -> str | None:
    if filename:
        lower = filename.lower()
        for ext in SUPPORTED_EXTENSIONS:
            if lower.endswith(ext):
                return ext
    mime_map = {
        "application/pdf": ".pdf",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "text/plain": ".txt",
        "text/markdown": ".md",
    }
    return mime_map.get(mime or "")


def _extract_text(data: bytes, ext: str) -> str:
    if ext in (".txt", ".md"):
        return data.decode("utf-8", errors="replace")

    if ext == ".pdf":
        import fitz  # type: ignore[import-untyped]

        doc = fitz.open(stream=data, filetype="pdf")
        pages = [doc[i].get_text() for i in range(doc.page_count)]
        return "\n\n".join(p for p in pages if p.strip())

    if ext == ".docx":
        import docx  # type: ignore[import-untyped]

        document = docx.Document(io.BytesIO(data))
        return "\n".join(p.text for p in document.paragraphs if p.text.strip())

    if ext == ".xlsx":
        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append("\t".join(cells))
            if rows:
                parts.append(f"### {sheet_name}\n" + "\n".join(rows))
        return "\n\n".join(parts)

    return ""


@router.message(lambda m: m.document is not None)
async def handle_document(message: Message, bot: Bot) -> None:
    """Handle document messages (.txt, .md, .pdf, .xlsx, .docx)."""
    if not message.document or not message.from_user:
        return

    doc = message.document
    mime = doc.mime_type or ""
    filename = doc.file_name or "document"
    ext = _detect_extension(filename, mime)

    if ext is None:
        await message.answer(
            f"⚠️ Формат не поддерживается: <code>{filename}</code>\n"
            "Поддерживаются: .txt .md .pdf .xlsx .docx"
        )
        return

    settings = get_settings()
    storage = VaultStorage(settings.vault_path)

    try:
        file = await bot.get_file(doc.file_id)
        if not file.file_path:
            await message.answer("❌ Не удалось скачать файл")
            return

        file_bytes_io = await bot.download_file(file.file_path)
        if not file_bytes_io:
            await message.answer("❌ Не удалось скачать файл")
            return

        data = file_bytes_io.read()
        text = _extract_text(data, ext)

        if not text.strip():
            await message.answer(f"⚠️ Файл пустой или не удалось извлечь текст: {filename}")
            return

        truncated = False
        if len(text) > MAX_TEXT_CHARS:
            text = text[:MAX_TEXT_CHARS]
            truncated = True

        content_parts = []
        if message.caption:
            content_parts.append(message.caption)
        content_parts.append(text)
        if truncated:
            content_parts.append(f"[…текст обрезан до {MAX_TEXT_CHARS} символов]")
        content = "\n\n".join(content_parts)

        timestamp = datetime.fromtimestamp(message.date.timestamp())
        storage.append_to_daily(content, timestamp, f"[doc: {filename}]")

        session = SessionStore(settings.vault_path)
        session.append(
            message.from_user.id,
            "document",
            filename=filename,
            text=content,
            msg_id=message.message_id,
        )

        caption_note = " + комментарий" if message.caption else ""
        trunc_note = " (обрезан)" if truncated else ""
        await message.answer(f"📄 ✓ Сохранено: {filename}{caption_note}{trunc_note}")
        logger.info("Document saved: %s (%d chars)", filename, len(text))

    except Exception:
        logger.exception("Error processing document: %s", filename)
        await message.answer(f"❌ Не удалось обработать файл: {filename}")
